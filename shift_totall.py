
#このコードはシフトはAIにてシフト内容をチェックすっるコードです。
# -*- coding: utf-8 -*-
from pathlib import Path
from typing import List, Optional, Union

#指示リスト作成
import openpyxl as pyxl
import os
import pandas as pd
import datetime

import jpholiday
import shutil
import time

from runtime_paths import resolve_shift_runtime_paths

USER = os.environ.get("USERNAME") or Path.home().name

SHOPCOUNT:int = 23


def SHIFT_CREATE(USER,y,m):
    #===============================================================================================
    runtime_paths = resolve_shift_runtime_paths(int(y), int(m))
    ONEDRIVE_PATH = runtime_paths.shift_excel_path
    COPYFILE_PATH = runtime_paths.copyfile_path
    TIMESCHEDULER_FILE = runtime_paths.timescheduler_path
    COPY_TIMESCHEDULER_FILE = runtime_paths.timescheduler_backup_path
    BugetFile = runtime_paths.budget_path
    DeliveryFile = runtime_paths.delivery_path
    COPYFILE_PATH.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy(ONEDRIVE_PATH,COPYFILE_PATH)
    #shutil.copy(COPY_TIMESCHEDULER_FILE,TIMESCHEDULER_FILE)
    # shutil.copy(BugetFile_OneDrive,BugetFile)
    # shutil.copy(DeliveryFile_OneDrive,DeliveryFile)
    time.sleep(5)


    PATH = COPYFILE_PATH

    WB = pyxl.load_workbook(PATH,data_only=True)

    SHEET_NAMES = ['本部【1】','所属登録',]

    WS_S  = WB[SHEET_NAMES[0]]

    WS_REG = WB[SHEET_NAMES[1]]

    SHOP_KEY = {'柏':'柏T',
                '千葉':'千葉T',
                '伊勢崎':'伊勢崎T',
                '富士見':'富士見T',
                'レイク':'レイクT',
                '海老名':'海老名T',
                'むさし':'むさしT',
                '平塚':'平塚T',
                '岡山':'岡山T',
                '大高':'大高T',
                '東郷町':'東郷T',
                '太田':'太田T',
                '水戸':'水戸T',
                'エキスポ':'EXPOT',
                '川崎':'川崎T',
                '新三郷':'新三郷T',
                '幕張':'幕張T',
                '各務原':'各務原T',
                '堺':'堺T',
                '四條畷':'四條畷T',
                '高崎' : '高崎T',
                '所沢':'所沢T',
                '梅田':'梅田T',
            
                }

    SHOP_KEY2 = {'FUN柏':'柏',
                'FUN千葉C-one':'千葉',
                'FUNスマーク伊勢崎':'伊勢崎',
                'FUNららぽーと富士見':'富士見',
                'FUNイオンレイクタウン':'レイク',
                'FUNららぽーと海老名':'海老名',
                'FUNイオンモールむさし村山':'むさし',
                'FUNららぽーと湘南平塚':'平塚',
                'FUNイオンモール岡山':'岡山',
                'FUNイオンモール大高':'大高',
                'FUNららぽーと愛知東郷':'東郷町',
                'FUNイオンモール太田':'太田',
                'FUNイオンモール水戸内原':'水戸',
                'FUNららぽーとEXPOCITY':'エキスポ',
                'FUNラゾーナ川崎プラザ':'川崎',
                'FUNららぽーと新三郷':'新三郷',
                'FUNイオンモール幕張新都心':'幕張',
                'FUNイオンモール各務原':'各務原',
                'FUNららぽーと堺':'堺',
                'FUNイオンモール四條畷':'四條畷',
                'FUNイオンモール高崎':'高崎',
                'FUNエミテラス所沢':'所沢',
                'FUN ＬＩＮＫＳ ＵＭＥＤＡ':'梅田',
                
                }

    SHOP_KEY_DIC = {
                '柏':'柏',
                '千葉':'千葉',
                '伊勢崎':'伊勢崎',
                '富士見':'富士見',
                'レイク':'レイク',
                '海老名':'海老名',
                'むさし':'むさし',
                '平塚':'平塚',
                '岡山':'岡山',
                '大高':'大高',
                '東郷町':'東郷町',
                '太田':'太田',
                '水戸':'水戸',
                'エキスポ':'エキスポ',
                '川崎':'川崎',
                '新三郷':'新三郷',
                '幕張':'幕張',
                '各務原':'各務原',
                '堺':'堺',
                '四條畷':'四條畷',
                '高崎':'高崎',
                '所沢':'所沢',
                '梅田':'梅田',
                }

    BUGET_KEY = {'柏':'H',
                '千葉':'I',
                '伊勢崎':'J',
                '富士見':'M',
                'レイク':'N',
                '海老名':'O',
                'むさし':'P',
                '平塚':'Q',
                '岡山':'R',
                '大高':'S',
                '東郷町':'T',
                '太田':'U',
                '水戸':'V',
                'エキスポ':'W',
                '川崎':'X',
                '新三郷':'Y',
                '幕張':'Z',
                '各務原':'AA',
                '堺':'AB',
                '四條畷':'K',
                '高崎':'L',
                '所沢':'AC',
                '梅田':'AD',
                
                }

    #---------------------------------------------------


    #＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝＝

    COL_H = ['Z','AE','AJ','AO','AT','AY','BD','BI','BN','BS','BX','CC','CH','CM','CR','CW','DB','DG','DL','DQ'] #本部COL
    #COL_T = ['AT','BB','BJ','BR','BZ','CH','CP','CX','DF','DN']#旧20240228
    COL_T = ['AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD']#新20250129

    #調整値
    row_add_H = 0
    row_add_T = -1

    #本部STAFFのシフトデータリストを作成
    STAFF_LIST_HONBU = []
    for i in range(20) :
        STAFF_CD = WS_REG['C' + str(253 + i)].value #社員CD
        
        if  STAFF_CD == None :
            print('None')
            
        else:    
            STAFF_NAME = WS_REG['E' + str(253 + i)].value #氏名
            STAFF_AFF = WS_REG['D' + str(253 + i)].value #所属
            STAFF_SALES = WS_REG['G' + str(253 + i)].value #販売力
            D1 = WS_S[COL_H[i] + str(13)].value 
            D2 = WS_S[COL_H[i] + str(16)].value 
            D3 = WS_S[COL_H[i] + str(19)].value 
            D4 = WS_S[COL_H[i] + str(22)].value 
            D5 = WS_S[COL_H[i] + str(25)].value 
            D6 = WS_S[COL_H[i] + str(28)].value 
            D7 = WS_S[COL_H[i] + str(31)].value 
            D8 = WS_S[COL_H[i] + str(34)].value 
            D9 = WS_S[COL_H[i] + str(37)].value 
            D10 = WS_S[COL_H[i] + str(40)].value 
            D11 = WS_S[COL_H[i] + str(43)].value 
            D12 = WS_S[COL_H[i] + str(46)].value 
            D13 = WS_S[COL_H[i] + str(49)].value 
            D14 = WS_S[COL_H[i] + str(52)].value 
            D15 = WS_S[COL_H[i] + str(55)].value 
            D16 = WS_S[COL_H[i] + str(58)].value 
            D17 = WS_S[COL_H[i] + str(61)].value 
            D18 = WS_S[COL_H[i] + str(64)].value 
            D19 = WS_S[COL_H[i] + str(67)].value 
            D20 = WS_S[COL_H[i] + str(70)].value 
            D21 = WS_S[COL_H[i] + str(73)].value 
            D22 = WS_S[COL_H[i] + str(76)].value 
            D23 = WS_S[COL_H[i] + str(79)].value 
            D24 = WS_S[COL_H[i] + str(82)].value 
            D25 = WS_S[COL_H[i] + str(85)].value 
            D26 = WS_S[COL_H[i] + str(88)].value 
            D27 = WS_S[COL_H[i] + str(91)].value 
            D28 = WS_S[COL_H[i] + str(94)].value 
            D29 = WS_S[COL_H[i] + str(97)].value 
            D30 = WS_S[COL_H[i] + str(100)].value 
            D31 = WS_S[COL_H[i] + str(103)].value 
            
            STAFF_DATA = pd.DataFrame([{'社員CD' :STAFF_CD,'氏名':STAFF_NAME,'所属':STAFF_AFF,'販売力':STAFF_SALES,
                                            'D1' : D1,
                                            'D2' : D2,
                                            'D3' : D3,
                                            'D4' : D4,
                                            'D5' : D5,
                                            'D6' : D6,
                                            'D7' : D7,
                                            'D8' : D8,
                                            'D9' : D9,
                                            'D10' : D10,
                                            'D11' : D11,
                                            'D12' : D12,
                                            'D13' : D13,
                                            'D14' : D14,
                                            'D15' : D15,
                                            'D16' : D16,
                                            'D17' : D17,
                                            'D18' : D18,
                                            'D19' : D19,
                                            'D20' : D20,
                                            'D21' : D21,
                                            'D22' : D22,
                                            'D23' : D23,
                                            'D24' : D24,
                                            'D25' : D25,
                                            'D26' : D26,
                                            'D27' : D27,
                                            'D28' : D28,
                                            'D29' : D29,
                                            'D30' : D30,
                                            'D31' : D31
                                        
                                        }])
            STAFF_LIST_HONBU.append(STAFF_DATA)
        
        
    CONCAT_STAFF_LIST_HONBU = pd.concat(STAFF_LIST_HONBU)    

    tenpo_counter = 0
    STAFF_LIST_TENPO = []
    for i_2 in range(SHOPCOUNT):
        #try :
        
        for i_3 in range(10):
            
            tenpo_name_full = WS_REG['A' + str(3 + i_3 + tenpo_counter)].value
            
            try:
                if tenpo_name_full == '長町' :
                    
                    print('No')
                elif tenpo_name_full == '船橋' :  
                    print('No')
                    
                else :
                    tenpo_name_key = SHOP_KEY_DIC[tenpo_name_full]
                    STAFF_CD = WS_REG['C' + str(3 + i_3 + tenpo_counter)].value
                    
                    if STAFF_CD == None :
                        print('None')
                        
                    else:
                        
                        STAFF_NAME = WS_REG['E' + str(3 + i_3 + tenpo_counter)].value
                        STAFF_AFF = WS_REG['D' + str(3 + i_3 + tenpo_counter)].value
                        STAFF_SALES = WS_REG['G' + str(3 + i_3 + tenpo_counter)].value
                        WS_T = WB[tenpo_name_key]
                        

                        D1 = WS_T[COL_T[i_3] + str(16 + row_add_T)].value 
                        D2 = WS_T[COL_T[i_3] + str(19 + row_add_T)].value 
                        D3 = WS_T[COL_T[i_3] + str(22 + row_add_T)].value 
                        D4 = WS_T[COL_T[i_3] + str(25 + row_add_T)].value 
                        D5 = WS_T[COL_T[i_3] + str(28 + row_add_T)].value 
                        D6 = WS_T[COL_T[i_3] + str(31 + row_add_T)].value 
                        D7 = WS_T[COL_T[i_3] + str(34 + row_add_T)].value 
                        D8 = WS_T[COL_T[i_3] + str(37 + row_add_T)].value 
                        D9 = WS_T[COL_T[i_3] + str(40 + row_add_T)].value 
                        D10 = WS_T[COL_T[i_3] + str(43 + row_add_T)].value 
                        D11 = WS_T[COL_T[i_3] + str(46 + row_add_T)].value 
                        D12 = WS_T[COL_T[i_3] + str(49 + row_add_T)].value 
                        D13 = WS_T[COL_T[i_3] + str(52 + row_add_T)].value 
                        D14 = WS_T[COL_T[i_3] + str(55 + row_add_T)].value 
                        D15 = WS_T[COL_T[i_3] + str(58 + row_add_T)].value 
                        D16 = WS_T[COL_T[i_3] + str(61 + row_add_T)].value 
                        D17 = WS_T[COL_T[i_3] + str(64 + row_add_T)].value 
                        D18 = WS_T[COL_T[i_3] + str(67 + row_add_T)].value 
                        D19 = WS_T[COL_T[i_3] + str(70 + row_add_T)].value 
                        D20 = WS_T[COL_T[i_3] + str(73 + row_add_T)].value 
                        D21 = WS_T[COL_T[i_3] + str(76 + row_add_T)].value 
                        D22 = WS_T[COL_T[i_3] + str(79 + row_add_T)].value 
                        D23 = WS_T[COL_T[i_3] + str(82 + row_add_T)].value 
                        D24 = WS_T[COL_T[i_3] + str(85 + row_add_T)].value 
                        D25 = WS_T[COL_T[i_3] + str(88 + row_add_T)].value 
                        D26 = WS_T[COL_T[i_3] + str(91 + row_add_T)].value 
                        D27 = WS_T[COL_T[i_3] + str(94 + row_add_T)].value 
                        D28 = WS_T[COL_T[i_3] + str(97 + row_add_T)].value 
                        D29 = WS_T[COL_T[i_3] + str(100 + row_add_T)].value 
                        D30 = WS_T[COL_T[i_3] + str(103 + row_add_T)].value 
                        D31 = WS_T[COL_T[i_3] + str(106 + row_add_T)].value 
                        
                        
                        STAFF_DATA = pd.DataFrame([{'店舗名':tenpo_name_full,'店舗Key' : tenpo_name_key,'社員CD' :STAFF_CD,'氏名':STAFF_NAME,'所属':STAFF_AFF,'販売力':STAFF_SALES,
                                                'D1' : D1,
                                                'D2' : D2,
                                                'D3' : D3,
                                                'D4' : D4,
                                                'D5' : D5,
                                                'D6' : D6,
                                                'D7' : D7,
                                                'D8' : D8,
                                                'D9' : D9,
                                                'D10' : D10,
                                                'D11' : D11,
                                                'D12' : D12,
                                                'D13' : D13,
                                                'D14' : D14,
                                                'D15' : D15,
                                                'D16' : D16,
                                                'D17' : D17,
                                                'D18' : D18,
                                                'D19' : D19,
                                                'D20' : D20,
                                                'D21' : D21,
                                                'D22' : D22,
                                                'D23' : D23,
                                                'D24' : D24,
                                                'D25' : D25,
                                                'D26' : D26,
                                                'D27' : D27,
                                                'D28' : D28,
                                                'D29' : D29,
                                                'D30' : D30,
                                                'D31' : D31

                                                }])
                        

                        STAFF_LIST_TENPO.append(STAFF_DATA)
            
            except:
                print("No Match")            

        tenpo_counter += 10


    CONCAT_STAFF_LIST_TENPO = pd.concat(STAFF_LIST_TENPO)
        
    return CONCAT_STAFF_LIST_HONBU,CONCAT_STAFF_LIST_TENPO,SHOP_KEY,SHOP_KEY2,BUGET_KEY
#===============================================================================================


#print(Concat_CHECKLIST)

#=============================================================================

tenpo_col_list = {
    
        "FUN柏":2,
        "FUN千葉C-one":9,
        "FUNスマーク伊勢崎":16,
        "FUNららぽーと富士見":37,
        "FUNイオンレイクタウン":44,
        "FUNららぽーと海老名":51,
        "FUNイオンモールむさし村山":58,
        "FUNららぽーと湘南平塚":65,
        "FUNイオンモール岡山":72,
        "FUNイオンモール大高":79,
        "FUNららぽーと愛知東郷":86,
        "FUNイオンモール太田":93,
        "FUNイオンモール水戸内原":100,
        "FUNららぽーとEXPOCITY":107,
        "FUNラゾーナ川崎プラザ":114,
        "FUNららぽーと新三郷":121,
        "FUNイオンモール幕張新都心":128,
        "FUNイオンモール各務原":135,
        "FUNららぽーと堺":142,
        'FUNイオンモール四條畷':23,
        'FUNイオンモール高崎':30,
        'FUNエミテラス所沢':149,
        'FUN ＬＩＮＫＳ ＵＭＥＤＡ':156,
    }

    #一致店舗の列調整
tenpo_pitch = {
'FUN柏':0,
'FUN千葉C-one':1,
'FUNスマーク伊勢崎':2,
'FUNららぽーと富士見':5,
'FUNイオンレイクタウン':6,
'FUNららぽーと海老名':7,
'FUNイオンモールむさし村山':8,
'FUNららぽーと湘南平塚':9,
'FUNイオンモール岡山':10,
'FUNイオンモール大高':11,
'FUNららぽーと愛知東郷':12,
'FUNイオンモール太田':13,
'FUNイオンモール水戸内原':14,
'FUNららぽーとEXPOCITY':15,
'FUNラゾーナ川崎プラザ':16,
'FUNららぽーと新三郷':17,
'FUNイオンモール幕張新都心':18,
'FUNイオンモール各務原':19,
'FUNららぽーと堺':20,
'FUNイオンモール四條畷':3,
'FUNイオンモール高崎':4,
'FUNエミテラス所沢':21,
'FUN ＬＩＮＫＳ ＵＭＥＤＡ':22,

}

#予算取得関数																			
def Buget(SHOP_KEYWORD,TargetDATE):
    
    col_dic = {

    "柏":[8,3],
    "千葉":[9,5],	
    "伊勢崎":[10,7],
    "四條畷":[11,9],
    "高崎":[12,11],
    "富士見":[13,13],
    "レイク":[14,15],
    "海老名":[15,17],
    "むさし":[16,19],
    "平塚":[17,21],
    "岡山":[18,23],
    "大高":[19,25],
    "東郷町":[20,27],
    "太田":[21,29],
    "水戸":[22,31],
    "エキスポ":[23,33],
    "川崎":[24,35],
    "新三郷":[25,37],
    "幕張":[26,39],
    "各務原":[27,41],
    "堺":[28,43],
    "所沢":[29,45],
    '梅田':[30,47],

    }

    buget_path = resolve_shift_runtime_paths(TargetDATE.year, TargetDATE.month).budget_path
    WB_buget = pyxl.load_workbook(buget_path,data_only=True)
    WS_buget = WB_buget["集計"]

    Start_Day = datetime.datetime(datetime.datetime.today().year,1,1)
    #Start_Day = datetime.datetime(datetime.datetime(2024,11,1).year,1,1)
    Target_Day = TargetDATE
    # Target_Day = datetime.datetime(2025,9,7)+ datetime.timedelta(days=1)

    Diff_Day = Target_Day - Start_Day
    stutas_row = 13 + Diff_Day.days

    Buget_Data = WS_buget.cell(stutas_row,col_dic[SHOP_KEYWORD][0]).value
    return Buget_Data
   
#納品数取得関数  
def Delivery(SHOP_KEYWORD):

    col_dic = {

    "柏":[8,3],
    "千葉":[9,5],	
    "伊勢崎":[10,7],
    "四條畷":[11,9],
    "高崎":[12,11],
    "富士見":[13,13],
    "レイク":[14,15],
    "海老名":[15,17],
    "むさし":[16,19],
    "平塚":[17,21],
    "岡山":[18,23],
    "大高":[19,25],
    "東郷町":[20,27],
    "太田":[21,29],
    "水戸":[22,31],
    "エキスポ":[23,33],
    "川崎":[24,35],
    "新三郷":[25,37],
    "幕張":[26,39],
    "各務原":[27,41],
    "堺":[28,43],
    "所沢":[29,45],
    "梅田":[30,47],

    }
    delivery_path = resolve_shift_runtime_paths(Target_Day2.year, Target_Day2.month).delivery_path
    WB_deli = pyxl.load_workbook(delivery_path,data_only=True)
    WS_deli = WB_deli["納品データ"]

    Start_Day2 = datetime.datetime(2022,1,1)
    #Target_Day2 = datetime.datetime.today()
    Target_Day2 = datetime.datetime(2024,11,1)

    Diff_Day2 = Target_Day2 - Start_Day2
    stutas_row2 = 4 + Diff_Day2.days

    Delivery_Data = WS_deli.cell(stutas_row2,col_dic[SHOP_KEYWORD][1]).value
    return Delivery_Data



# データベースに接続する
driver = "SQL Server"
server ="FUN-PC132"
database = 'TimeData'#時間帯売上実績
trusted_connection = "yes"
# conn = pyodbc.connect('DRIVER='+driver+';SERVER='+server+';DATABASE='+database+';POST=1433;Trusted_Connection='+trusted_connection+';')


# cursor = conn.cursor()

# def SELECT_TODAY(y,m,d):
#     cursor.execute('SELECT * FROM  SalesData WHERE Year =' + str(y) + ' AND Month = ' + str(m) + ' AND Day = ' + str(d) + '')


TimeSalesList = []


def SELECT(shop,year,month,day):
    #select_day = str(year) + "-" + str(month).zfill(2) +"-" + str(day).zfill(2) + " 00:00:00"
    shop = shop


    # cursor.execute("SELECT * FROM  TimeData WHERE Shop ='" + str(shop) + "' AND Year =" + str(year) + " AND Month = " + str(month) + " AND Day = " + str(day) + "")
    # #print(shop)
    
    # try:
    #     for row in cursor:
    #         #print(row)
    #         # row
    #         timedata = pd.DataFrame({"店舗":[row[0]],"曜日特性":[row[5]],
    #                                 "t8":[row[7]],
    #                                 "t9":[row[8]],
    #                                 "t10":[row[9]],
    #                                 "t11":[row[10]],
    #                                 "t12":[row[11]],
    #                                 "t13":[row[12]],
    #                                 "t14":[row[13]],
    #                                 "t15":[row[14]],
    #                                 "t16":[row[15]],
    #                                 "t17":[row[16]],
    #                                 "t18":[row[17]],
    #                                 "t19":[row[18]],
    #                                 "t20":[row[19]],
    #                                 "t21":[row[20]],
    #                                 "t22":[row[21]],

    #                                 })
            
    #         TimeSalesList.append(timedata)
            
    # except:
    #     timedata = pd.DataFrame({"店舗":[shop],"曜日特性":['平日        '],
    #                                 "t8":[0],
    #                                 "t9":[0],
    #                                 "t10":[0],
    #                                 "t11":[0],
    #                                 "t12":[0],
    #                                 "t13":[0],
    #                                 "t14":[0],
    #                                 "t15":[0],
    #                                 "t16":[0],
    #                                 "t17":[0],
    #                                 "t18":[0],
    #                                 "t19":[0],
    #                                 "t20":[0],
    #                                 "t21":[0],
    #                                 "t22":[0],

    #                                 })
    #     #print(timedata)
            
        
    #     TimeSalesList.append(timedata)
#=============================================================================



#履歴書フォーマット作成
def print_string(filename,TargetDATE,SHIFTDATA1,SHIFTDATA2):
        SHOP_KEY2 = {'FUN柏':'柏',
                'FUN千葉C-one':'千葉',
                'FUNスマーク伊勢崎':'伊勢崎',
                'FUNららぽーと富士見':'富士見',
                'FUNイオンレイクタウン':'レイク',
                'FUNららぽーと海老名':'海老名',
                'FUNイオンモールむさし村山':'むさし',
                'FUNららぽーと湘南平塚':'平塚',
                'FUNイオンモール岡山':'岡山',
                'FUNイオンモール大高':'大高',
                'FUNららぽーと愛知東郷':'東郷町',
                'FUNイオンモール太田':'太田',
                'FUNイオンモール水戸内原':'水戸',
                'FUNららぽーとEXPOCITY':'エキスポ',
                'FUNラゾーナ川崎プラザ':'川崎',
                'FUNららぽーと新三郷':'新三郷',
                'FUNイオンモール幕張新都心':'幕張',
                'FUNイオンモール各務原':'各務原',
                'FUNららぽーと堺':'堺',
                'FUNイオンモール四條畷':'四條畷',
                'FUNイオンモール高崎':'高崎',
                'FUNエミテラス所沢':'所沢',
                'FUN ＬＩＮＫＳ ＵＭＥＤＡ':'梅田',
                
                }
        print("スタート",filename)
    #for shop_name in tenpo_pitch :
        shop_key = shop_name
        print("shop_key⇒",shop_key)
        Y = TargetDATE.year
        M = TargetDATE.month
        D = TargetDATE.day
        
        CONCAT_STAFF_LIST_TENPO = SHIFTDATA1
        CONCAT_STAFF_LIST_HONBU = SHIFTDATA2
        #3週間前のデータを参照
        for i in range(28):
            diff_day = datetime.timedelta(days=i)
            DateTime = TargetDATE - diff_day
            year = int(DateTime.year)
            # month = 7#int(DateTime.month)
            # day = 30 - i #int(DateTime.day)    
            month = int(DateTime.month)
            day = int(DateTime.day)   
            
            SELECT(shop=shop_key,year=year,month=month,day=day)
            
        #Concat_TimeSaleList = pd.concat(TimeSalesList)

        #================================================================ 
        #新コード
        # try :       
        #     Today_Delivery = P_DATA(shop_key)
        # except :
        #     Today_Delivery = 0
            
        week = ['月','火','水','木','金','土','日']
        today_1 = TargetDATE
        # today_1 = datetime.date(2025,9,7) + datetime.timedelta(days=1)
        f_day = today_1.strftime("%Y%m%d")#西暦表記
        y_day = today_1.strftime("%Y")#西暦表記
        m_day = today_1.strftime("%m")#月
        d_day =today_1.strftime("%d")#日
        w_day = today_1.weekday()
        
        print("check",d_day)

        DOW = week[w_day]
        if DOW == None:
            DOW = 0

        if (DOW == "土") or (DOW == "日"):     
            DOW_Type =  "土日祝       "   

        else:
            Day_holi = jpholiday.is_holiday(datetime.date(int(y_day),int(m_day),int(d_day)))    
            
            print(Day_holi , "tetst")
            if Day_holi == True:
                DOW_Type =  "土日祝       " 
            else:
                DOW_Type =  "平日        "   

        #for shop_name2 in tenpo_pitch :
        print(DOW_Type)
        #SELECT_DATA = Concat_TimeSaleList[(Concat_TimeSaleList["店舗"] == str(shop_name)) & (Concat_TimeSaleList["曜日特性"] == str(DOW_Type))]
        
        # print("一致数",len(SELECT_DATA.query('t12 > 0')))
        # print("合計金額",sum(SELECT_DATA['t12'].values))
        ToDay_Key = int(d_day)
        Help_List =[]        
        var_List = []
        for i_4 in CONCAT_STAFF_LIST_TENPO.values :
            
            #print(row,i_4)
        
            ToDay_Shift = i_4[5 + ToDay_Key]
            
            #print(ToDay_Shift)
            if ToDay_Shift == None :
                print('NoData')
                
            else :
                
                # ensure shift is treated as string before membership check
                try:
                    if SHOP_KEY2[shop_name] in str(ToDay_Shift):
                        
                        HELP_DATA = pd.DataFrame([{'ヘルプ店舗':SHOP_KEY2[shop_name],'所属店舗':i_4[0],'所属KEY':i_4[1],'社員CD':i_4[2],'氏名':i_4[3],'役職':i_4[4],'販売力':i_4[5],'シフト':i_4[5 + ToDay_Key]}])
                        Help_List.append(HELP_DATA)
                except Exception:
                    # ignore non-iterable/unexpected values
                    pass
                    
        for i_5 in CONCAT_STAFF_LIST_HONBU.values :
            ToDay_Shift = i_5[3 + ToDay_Key]
            if ToDay_Shift == None :
                print('NoData')
                
            else :
                
                # ensure shift is treated as string before membership check
                try:
                    if SHOP_KEY2[shop_name] in str(ToDay_Shift):
                        
                        HELP_DATA = pd.DataFrame([{'ヘルプ店舗':SHOP_KEY2[shop_name],'所属店舗':'本部','所属KEY':'本部','社員CD':i_5[0],'氏名':i_5[1],'役職':i_5[2],'販売力':i_5[3],'シフト':i_5[3 + ToDay_Key]}])
                        Help_List.append(HELP_DATA)
                except Exception:
                    pass
        
        try:
            CONCAT_HELP_LIST = pd.concat(Help_List)
            #print(CONCAT_HELP_LIST)
            
        except :
            # use an empty DataFrame with expected columns to avoid unbound-variable / indexing issues
            CONCAT_HELP_LIST = pd.DataFrame(columns=['ヘルプ店舗','所属店舗','所属KEY','社員CD','氏名','役職','販売力','シフト'])
            #print("none")
        #print(CONCAT_HELP_LIST) 
        Help_COLLIST = ['K','L','M','N']   
    
                        
        #================================================================
        #ここまで
                    
        # for time_zone in range(8,23):
        #     target = len(SELECT_DATA.query('t' + str(time_zone) + ' > 0'))


        #     if target == 0:
        #         av_data = pd.DataFrame({"時間帯":["t" + str(time_zone)],"売上":[0]})
        #         var_List.append(av_data)
                
        #     elif time_zone < 10:
        #         av_data = pd.DataFrame({"時間帯":["t" + str(time_zone)],"売上":[0]})
        #         var_List.append(av_data)
                
                
        #     elif time_zone > 19:
        #         av_data = pd.DataFrame({"時間帯":["t" + str(time_zone)],"売上":[0]})
        #         var_List.append(av_data)  
                
                
                
        #     else:
        #         targetsum = sum(SELECT_DATA['t' + str(time_zone)].values)
        #         av_data = pd.DataFrame({"時間帯":["t" + str(time_zone)],"売上":[ targetsum / target]})
        #         var_List.append(av_data)
                
        # concat_var_List = pd.concat(var_List)   
        # sum_sales = sum(concat_var_List["売上"].values)
        
        # ratio_list = []
        # for row_data in concat_var_List.values:
        #     ratio = (row_data[1]/sum_sales)*100
        #     ratio2 = "{: .1f}".format(ratio)
        #     ratio_list.append(ratio2)
 
        staff_count = ""
        # 店別詳細
        # tableを作成
    
        #P処理目標時間
        # mini_standerd = 10
        # print("納品数=>",Today_Delivery)
        # if Today_Delivery == 0 :
            
        #    p_time_str = ""           
            
        # else :
        #     p_time = int(Today_Delivery) * mini_standerd/60 
        #     p_time_hours = int(p_time)
        #     p_time_minutes = int(round((p_time - p_time_hours) * 60 / 15) * 15)
        #     if p_time_minutes == 60:
        #         p_time_hours += 1
        #         p_time_minutes = 0
        #     p_time_str = f"{p_time_hours}時間{p_time_minutes}分"   
        
        
          
        staff_post = []
        staff_name = []
        staff_sales = []
        staff_shift = []
               
        MATCH_DATA_ToShop = CONCAT_STAFF_LIST_TENPO[CONCAT_STAFF_LIST_TENPO['店舗Key'] == SHOP_KEY2[shop_name]].values
        row_counter3 = 0
        
        #for toshopshift_data in MATCH_DATA_ToShop:
        for index1 in range(10):

            try:
                #print(MATCH_DATA_ToShop[index1])
                staff_post.append(MATCH_DATA_ToShop[index1][4])
                staff_name.append(MATCH_DATA_ToShop[index1][3])
                try:
                    staff_sales.append('{: .1f}'.format(MATCH_DATA_ToShop[index1][5]))
                    
                except:
                    staff_sales.append("")
                        
                staff_shift.append(MATCH_DATA_ToShop[index1][5 + ToDay_Key])
            
                row_counter3 += 1
                
            except IndexError:
                staff_post.append("")
                staff_name.append("")
                staff_sales.append("")
                staff_shift.append("")
                
                
                
                    
        H_staff_post = []
        H_staff_name = []
        H_staff_sales = []
        H_staff_shift = []    
            
        try :    
            MATCH_DATA = CONCAT_HELP_LIST[CONCAT_HELP_LIST['ヘルプ店舗'] == SHOP_KEY2[shop_name]]
            #if len(MATCH_DATA) >= 1:
            H_staff = MATCH_DATA[MATCH_DATA['所属KEY'] == '本部'].values
            #try :
            row_counter1 = 0
            #for output_data in H_staff :
            for index2 in range(5):
                
                try:
                    H_staff_post.append(H_staff[index2][5])
                    H_staff_name.append(H_staff[index2][4])
                    try :
                        H_staff_sales.append('{: .1f}'.format(H_staff[index2][6]))
                        
                    except:
                        H_staff_sales.append('')
                        
                    H_staff_shift.append(H_staff[index2][7])
                
                    row_counter1 += 1
                    
                    
                except IndexError:
                    
                    H_staff_post.append("")
                    H_staff_name.append("")
                    H_staff_sales.append("")
                    H_staff_shift.append("")
                    row_counter1 += 1
                    
        except :
            H_staff_post = ["","","","",""]
            H_staff_name = ["","","","",""]
            H_staff_sales = ["","","","",""]
            H_staff_shift = ["","","","",""]  
            
        T_staff_post = []
        T_staff_name = []
        T_staff_sales = []
        T_staff_shift = []   
                
            
        try:
            T_staff = MATCH_DATA[MATCH_DATA['所属KEY'] != '本部'].values
            row_counter2 = 0
            
            #for output_data2 in T_staff :
            for index3 in range(5):
                
                try:
                    T_staff_post.append(T_staff[index3][5])
                    T_staff_name.append(T_staff[index3][4])
                    try :
                        T_staff_sales.append('{: .1f}'.format(T_staff[index3][6]))
                    except :
                        T_staff_sales.append('')
                            
                        
                    T_staff_shift.append(T_staff[index3][7])
                
                    row_counter2 += 1
                    
                except IndexError :
                    
                    T_staff_post.append("")
                    T_staff_name.append("")
                    T_staff_sales.append("")
                    T_staff_shift.append("")
                    row_counter2 += 1
                    
        except :
            T_staff_post = ["","","","",""]
            T_staff_name = ["","","","",""]
            T_staff_sales = ["","","","",""]
            T_staff_shift = ["","","","",""] 

        # print(staff_post,staff_name,staff_sales,staff_shift)           
        # print(T_staff_post,T_staff_name,T_staff_sales,T_staff_shift)
        # print(H_staff_post,H_staff_name,H_staff_sales,H_staff_shift)
            # 3つのリストを1つのDataFrameにまとめる
        staff_data = pd.DataFrame({
            "役職": staff_post,
            "氏名": staff_name,
            "販売力": staff_sales,
            "シフト": staff_shift
        })

        h_staff_data = pd.DataFrame({
            "役職": H_staff_post,
            "氏名": H_staff_name,
            "販売力": H_staff_sales,
            "シフト": H_staff_shift
        })

        t_staff_data = pd.DataFrame({
            "役職": T_staff_post,
            "氏名": T_staff_name,
            "販売力": T_staff_sales,
            "シフト": T_staff_shift
        })

        # 全体をまとめる
        all_staff_data = pd.concat([staff_data, h_staff_data, t_staff_data], ignore_index=True)
        
        # 左端に列を追加
        all_staff_data.insert(0, "店舗", shop_key)
        all_staff_data.insert(0, "日", d_day)
        all_staff_data.insert(0, "月", m_day)
        all_staff_data.insert(0, "年度", y_day)

        
        return all_staff_data

    
            
        

# エントリーポイント
if __name__ == "__main__":
    all_data_frames = []
    SELECT_FILE_DATA = SHIFT_CREATE(USER,2026,3)

    for shop_name in tenpo_pitch:
        for DATE_Counter in range(0,31):
            today = datetime.datetime(2026,3,1 + DATE_Counter)
            df = print_string(shop_name, today,SELECT_FILE_DATA[1],SELECT_FILE_DATA[0])
            all_data_frames.append(df)

    final_df = pd.concat(all_data_frames, ignore_index=True)
    print(final_df)
    

